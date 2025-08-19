"""
Excel Report Generator for IT Dashboard
Creates Excel reports with charts from GitHub data
"""
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
import json
import logging

class ITDashboardGenerator:
    def __init__(self, config: dict):
        self.config = config
        self.logger = logging.getLogger(__name__)
        plt.style.use(['default'])
        sns.set_palette("husl")

    def create_summary_charts(self, data: pd.DataFrame, output_dir: Path) -> dict:
        chart_files = {}
        plt.rcParams.update({'figure.dpi': 300, 'savefig.dpi': 300, 'font.size': 10})

        # 1. Stars Comparison
        fig, ax = plt.subplots(figsize=(10, 6))
        data_sorted = data.sort_values('stars', ascending=False)
        bars = ax.bar(range(len(data_sorted)), data_sorted['stars'])
        ax.set_xlabel('Repository')
        ax.set_ylabel('Stars Count')
        ax.set_title('GitHub Repository Stars Comparison', fontsize=14, fontweight='bold')
        ax.set_xticks(range(len(data_sorted)))
        ax.set_xticklabels(data_sorted['repository'].str.split('/').str[1], rotation=45, ha='right')
        for i, bar in enumerate(bars):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height, f'{int(height):,}', ha='center', va='bottom')
        plt.tight_layout()
        stars_chart = output_dir / 'stars_comparison.png'
        plt.savefig(stars_chart, bbox_inches='tight', facecolor='white')
        plt.close()
        chart_files['stars'] = stars_chart

        # 2. Open Issues vs PRs
        fig, ax = plt.subplots(figsize=(10, 6))
        x = range(len(data))
        width = 0.35
        ax.bar([i - width/2 for i in x], data['open_issues'], width, label='Open Issues', alpha=0.8)
        ax.bar([i + width/2 for i in x], data['open_prs'], width, label='Open PRs', alpha=0.8)
        ax.set_xlabel('Repository')
        ax.set_ylabel('Count')
        ax.set_title('Open Issues vs Pull Requests', fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(data['repository'].str.split('/').str[1], rotation=45, ha='right')
        ax.legend()
        plt.tight_layout()
        issues_prs_chart = output_dir / 'issues_vs_prs.png'
        plt.savefig(issues_prs_chart, bbox_inches='tight', facecolor='white')
        plt.close()
        chart_files['issues_prs'] = issues_prs_chart

        # 3. Size Distribution
        fig, ax = plt.subplots(figsize=(8, 8))
        sizes = data['size_kb'] / 1024
        labels = data['repository'].str.split('/').str[1]
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        ax.set_title('Repository Size Distribution (MB)', fontsize=14, fontweight='bold')
        plt.setp(autotexts, size=8, weight="bold")
        plt.setp(texts, size=9)
        pie_chart = output_dir / 'size_distribution.png'
        plt.savefig(pie_chart, bbox_inches='tight', facecolor='white')
        plt.close()
        chart_files['size_dist'] = pie_chart

        # 4. Activity Metrics Heatmap
        fig, ax = plt.subplots(figsize=(10, 6))
        metrics_data = data[['repository', 'stars', 'forks', 'open_issues', 'open_prs']].copy()
        metrics_data['repository'] = metrics_data['repository'].str.split('/').str[1]
        metrics_data = metrics_data.set_index('repository')
        normalized_data = (metrics_data - metrics_data.min()) / (metrics_data.max() - metrics_data.min())
        sns.heatmap(normalized_data.T, annot=True, fmt='.2f', cmap='YlOrRd', cbar_kws={'label': 'Normalized Score'})
        ax.set_title('Repository Activity Metrics (Normalized)', fontsize=14, fontweight='bold')
        ax.set_xlabel('Repository')
        ax.set_ylabel('Metrics')
        plt.tight_layout()
        heatmap_chart = output_dir / 'activity_heatmap.png'
        plt.savefig(heatmap_chart, bbox_inches='tight', facecolor='white')
        plt.close()
        chart_files['heatmap'] = heatmap_chart

        return chart_files

    def style_worksheet(self, ws, title: str):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        ws.insert_rows(1)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells(f'A1:{ws.max_column}1')

    def create_dashboard_report(self, data: pd.DataFrame) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = self.config["report"]["template_name"].format(date=timestamp)
        output_dir = Path(self.config["report"]["output_dir"])
        output_path = Path("reports") / report_name
        output_path.parent.mkdir(parents=True, exist_ok=True)
        charts_dir = output_path.parent / "charts" / timestamp
        charts_dir.mkdir(parents=True, exist_ok=True)
        chart_files = self.create_summary_charts(data, charts_dir)
        wb = Workbook()
        wb.remove(wb.active)
        summary_ws = wb.create_sheet("Dashboard Summary")
        summary_data = [
            ["Metric", "Value"],
            ["Total Repositories", len(data)],
            ["Total Stars", data['stars'].sum()],
            ["Total Forks", data['forks'].sum()],
            ["Total Open Issues", data['open_issues'].sum()],
            ["Total Open PRs", data['open_prs'].sum()],
            ["Average Repository Size (MB)", round(data['size_kb'].mean() / 1024, 2)],
            ["Most Popular Language", data['language'].mode().iloc[0] if not data['language'].mode().empty else "N/A"],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        for row in summary_data:
            summary_ws.append(row)
        self.style_worksheet(summary_ws, "IT Dashboard Summary")
        detail_ws = wb.create_sheet("Repository Details")
        for r in dataframe_to_rows(data, index=False, header=True):
            detail_ws.append(r)
        self.style_worksheet(detail_ws, "Detailed Repository Data")
        charts_ws = wb.create_sheet("Visual Analytics")
        row_position = 2
        for chart_name, chart_path in chart_files.items():
            try:
                img = Image(str(chart_path))
                img.width = 600
                img.height = 400
                charts_ws.add_image(img, f'A{row_position}')
                row_position += 22
            except Exception as e:
                self.logger.warning(f"Could not insert chart {chart_name}: {e}")
        charts_ws['A1'] = "Visual Analytics Dashboard"
        charts_ws['A1'].font = Font(bold=True, size=16, color="366092")
        wb.save(output_path)
        self.logger.info(f"✓ Excel dashboard saved: {output_path}")
        return str(output_path)

    def generate_daily_report(self, data: pd.DataFrame) -> str:
        try:
            report_path = self.create_dashboard_report(data)
            self.logger.info("Dashboard Report Summary:")
            self.logger.info(f"  - Repositories analyzed: {len(data)}")
            self.logger.info(f"  - Total open issues: {data['open_issues'].sum()}")
            self.logger.info(f"  - Total open PRs: {data['open_prs'].sum()}")
            self.logger.info(f"  - Most starred repo: {data.loc[data['stars'].idxmax(), 'repository']}")
            return report_path
        except Exception as e:
            self.logger.error(f"Error generating dashboard report: {e}")
            raise

def generate_excel_report(data: pd.DataFrame, config_path: str) -> str:
    with open(config_path, 'r') as f:
        config = json.load(f)
    logging.basicConfig(
        level=getattr(logging, config["logging"]["level"]),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    generator = ITDashboardGenerator(config)
    return generator.generate_daily_report(data)
